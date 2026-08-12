import os
import io
import re
import base64
import zipfile
import threading
from datetime import datetime, timedelta
from flask import Flask, jsonify, request, render_template, send_file
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
CORS(app)

DATABASE_URL = os.environ.get("DATABASE_URL", "")
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

if DATABASE_URL:
    app.config["SQLALCHEMY_DATABASE_URI"] = DATABASE_URL
else:
    DB_PATH = os.path.join(os.path.dirname(__file__), "reportes.db")
    app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{DB_PATH}"

app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
db = SQLAlchemy(app)

class Reporte(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    tienda     = db.Column(db.String(200))
    producto   = db.Column(db.String(200))
    comentario = db.Column(db.Text)
    foto       = db.Column(db.String(300))
    foto_b64   = db.Column(db.Text)
    usuario    = db.Column(db.String(100))
    fecha      = db.Column(db.String(20))
    semana     = db.Column(db.String(20), default="")  # semana a la que pertenece

class Inventario(db.Model):
    id       = db.Column(db.Integer, primary_key=True)
    tienda   = db.Column(db.String(200), index=True)
    producto = db.Column(db.String(200))
    cantidad = db.Column(db.Integer, default=0)

class Diferencia(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    semana     = db.Column(db.String(20), index=True)
    tienda     = db.Column(db.String(200))
    producto   = db.Column(db.String(200))
    cantidad   = db.Column(db.Integer, default=0)
    estado     = db.Column(db.String(50))   # SIN_FOTO / CON_JUSTIFICACION / OK
    comentario = db.Column(db.Text)

class Usuario(db.Model):
    id       = db.Column(db.Integer, primary_key=True)
    nombre   = db.Column(db.String(150))
    usuario  = db.Column(db.String(100), unique=True)
    password = db.Column(db.String(100))
    rol      = db.Column(db.String(20), default="display")

class Config(db.Model):
    clave = db.Column(db.String(50), primary_key=True)
    valor = db.Column(db.String(100))

class ValidacionIA(db.Model):
    __tablename__ = "validacion_ia"
    id           = db.Column(db.Integer, primary_key=True)
    reporte_id   = db.Column(db.Integer, db.ForeignKey("reporte.id"), nullable=True)
    semana       = db.Column(db.String(20), index=True)
    tienda       = db.Column(db.String(200))
    producto     = db.Column(db.String(200))
    marca        = db.Column(db.String(200))
    # PENDIENTE / PROCESANDO / APROBADO / RECHAZADO / REVISAR / ERROR
    estado       = db.Column(db.String(20), default="PENDIENTE", index=True)
    confianza    = db.Column(db.String(10))   # alta / media / baja
    motivo       = db.Column(db.Text)         # explicación corta de Claude
    tokens_input = db.Column(db.Integer, default=0)
    tokens_output= db.Column(db.Integer, default=0)
    costo_usd    = db.Column(db.Float, default=0.0)
    intentos     = db.Column(db.Integer, default=0)
    creado_en    = db.Column(db.String(30))
    procesado_en = db.Column(db.String(30))

USUARIOS_INICIALES = [
    {"nombre": "Display 1", "usuario": "display1", "password": "1234",     "rol": "display"},
    {"nombre": "Supervisor", "usuario": "admin",    "password": "admin123", "rol": "supervisor"}
]

def _migrar_columnas():
    # db.create_all() no altera tablas existentes; agrega columnas nuevas a mano.
    with db.engine.connect() as conn:
        try:
            if db.engine.dialect.name == "postgresql":
                conn.execute(db.text("ALTER TABLE reporte ADD COLUMN IF NOT EXISTS foto_b64 TEXT"))
            else:
                conn.execute(db.text("ALTER TABLE reporte ADD COLUMN foto_b64 TEXT"))
            conn.commit()
        except Exception:
            conn.rollback()

        # Migrar tabla validacion_ia si no existe (db.create_all la crea, pero por si acaso)
        try:
            if db.engine.dialect.name == "postgresql":
                conn.execute(db.text("""
                    CREATE TABLE IF NOT EXISTS validacion_ia (
                        id SERIAL PRIMARY KEY,
                        reporte_id INTEGER REFERENCES reporte(id) ON DELETE SET NULL,
                        semana VARCHAR(20),
                        tienda VARCHAR(200),
                        producto VARCHAR(200),
                        marca VARCHAR(200),
                        estado VARCHAR(20) DEFAULT 'PENDIENTE',
                        confianza VARCHAR(10),
                        motivo TEXT,
                        tokens_input INTEGER DEFAULT 0,
                        tokens_output INTEGER DEFAULT 0,
                        costo_usd FLOAT DEFAULT 0,
                        intentos INTEGER DEFAULT 0,
                        creado_en VARCHAR(30),
                        procesado_en VARCHAR(30)
                    )
                """))
            conn.commit()
        except Exception:
            conn.rollback()

def _seed_usuarios():
    if Usuario.query.count() == 0:
        for u in USUARIOS_INICIALES:
            db.session.add(Usuario(**u))
        db.session.commit()

with app.app_context():
    db.create_all()
    _migrar_columnas()
    _seed_usuarios()

SYNC_KEY = os.environ.get("SYNC_KEY", "cpfr2024")

FOTOS_DIR = os.path.join(os.path.dirname(__file__), '..', 'fotos')
os.makedirs(FOTOS_DIR, exist_ok=True)

def semana_actual():
    cfg = Config.query.get("semana_override")
    if cfg and cfg.valor:
        return cfg.valor
    return datetime.now().strftime("%Y-S%V")

MESES_ES = ["ene", "feb", "mar", "abr", "may", "jun", "jul", "ago", "sep", "oct", "nov", "dic"]

def rango_de_codigo(semana_code):
    """Convierte '2026-S30' al rango de fechas legible y para nombre de archivo."""
    try:
        parts = semana_code.split("-S")
        year, week = int(parts[0]), int(parts[1])
        lunes  = datetime.fromisocalendar(year, week, 1)
        domingo = lunes + timedelta(days=6)
    except Exception:
        now = datetime.now()
        lunes = now - timedelta(days=now.weekday())
        domingo = lunes + timedelta(days=6)
    mes_lunes   = MESES_ES[lunes.month - 1]
    mes_domingo = MESES_ES[domingo.month - 1]
    if lunes.month == domingo.month:
        archivo = f"{lunes.day:02d}-{domingo.day:02d}_{mes_domingo}_{domingo.year}"
        legible = f"{lunes.day} – {domingo.day} {mes_domingo} {domingo.year}"
    else:
        archivo = f"{lunes.day:02d}_{mes_lunes}-{domingo.day:02d}_{mes_domingo}_{domingo.year}"
        legible = f"{lunes.day} {mes_lunes} – {domingo.day} {mes_domingo} {domingo.year}"
    return archivo, legible

def rango_semana_actual():
    return rango_de_codigo(semana_actual())

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/ping")
def ping():
    return jsonify({"ok": True})

@app.route("/api/login", methods=["POST"])
def login():
    data     = request.json
    usuario  = data.get("usuario", "").strip()
    password = data.get("password", "").strip()
    u = Usuario.query.filter_by(usuario=usuario, password=password).first()
    if u:
        return jsonify({"ok": True, "id": u.id, "nombre": u.nombre, "rol": u.rol})
    return jsonify({"ok": False, "error": "Usuario o contraseña incorrectos"}), 401

@app.route("/api/semana-activa")
def get_semana_activa():
    codigo = semana_actual()
    _, legible = rango_de_codigo(codigo)
    return jsonify({"semana_activa": codigo, "semana_legible": legible})

@app.route("/api/set-semana", methods=["POST"])
def set_semana():
    data = request.json or {}
    semana = (data.get("semana") or "").strip()
    cfg = Config.query.get("semana_override")
    if semana:
        if not cfg:
            cfg = Config(clave="semana_override", valor=semana)
            db.session.add(cfg)
        else:
            cfg.valor = semana
    else:
        if cfg:
            cfg.valor = ""
    db.session.commit()
    return jsonify({"ok": True, "semana_activa": semana_actual()})

def _es_supervisor(data):
    return (data or {}).get("solicitante_rol") == "supervisor"

@app.route("/api/usuarios")
def listar_usuarios():
    rows = Usuario.query.order_by(Usuario.rol.desc(), Usuario.nombre).all()
    return jsonify([{"id": u.id, "nombre": u.nombre, "usuario": u.usuario, "password": u.password, "rol": u.rol} for u in rows])

@app.route("/api/usuarios", methods=["POST"])
def crear_usuario():
    data = request.json
    if not _es_supervisor(data):
        return jsonify({"ok": False, "error": "No autorizado"}), 403
    nombre   = data.get("nombre", "").strip()
    usuario  = data.get("usuario", "").strip()
    password = data.get("password", "").strip()
    if not (nombre and usuario and password):
        return jsonify({"ok": False, "error": "Faltan datos"}), 400
    if Usuario.query.filter_by(usuario=usuario).first():
        return jsonify({"ok": False, "error": "Ese usuario ya existe"}), 400
    u = Usuario(nombre=nombre, usuario=usuario, password=password, rol="display")
    db.session.add(u)
    db.session.commit()
    return jsonify({"ok": True, "id": u.id})

@app.route("/api/usuarios/<int:usuario_id>", methods=["DELETE"])
def eliminar_usuario(usuario_id):
    if not _es_supervisor(request.json):
        return jsonify({"ok": False, "error": "No autorizado"}), 403
    u = Usuario.query.get_or_404(usuario_id)
    if u.rol == "supervisor":
        return jsonify({"ok": False, "error": "No se puede eliminar al supervisor"}), 400
    db.session.delete(u)
    db.session.commit()
    return jsonify({"ok": True})

@app.route("/api/tiendas")
def tiendas():
    rows = db.session.query(Inventario.tienda).distinct().order_by(Inventario.tienda).all()
    return jsonify([r.tienda for r in rows])

@app.route("/api/tiendas-reportadas")
def tiendas_reportadas():
    rows = db.session.query(Reporte.tienda).filter_by(semana=semana_actual()).distinct().all()
    return jsonify([r.tienda for r in rows])

@app.route("/api/productos/<tienda>")
def productos(tienda):
    rows = Inventario.query.filter_by(tienda=tienda).order_by(Inventario.cantidad.desc()).all()
    return jsonify([{"nombre": r.producto, "cantidad": r.cantidad} for r in rows])

@app.route("/api/sync", methods=["POST"])
def sync_inventario():
    data = request.json
    if data.get("key") != SYNC_KEY:
        return jsonify({"ok": False, "error": "No autorizado"}), 403
    inventario = data.get("inventario", {})
    Inventario.query.delete()
    total = 0
    for tienda, prods in inventario.items():
        for p in prods:
            db.session.add(Inventario(tienda=tienda, producto=p["nombre"], cantidad=p["cantidad"]))
            total += 1
    db.session.commit()
    return jsonify({"ok": True, "productos": total})

@app.route("/api/reporte", methods=["POST"])
def guardar_reporte():
    data        = request.json
    tienda      = data.get("tienda", "")
    producto    = data.get("producto", "")
    comentario  = data.get("comentario", "")
    foto_b64    = data.get("foto", "")
    usuario     = data.get("usuario", "")
    fecha       = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_foto = ""
    if foto_b64:
        nombre_foto = f"{fecha}_{tienda}_{producto[:20]}.jpg".replace(" ", "_")
        with open(os.path.join(FOTOS_DIR, nombre_foto), "wb") as f:
            f.write(base64.b64decode(foto_b64))
    rep = Reporte(tienda=tienda, producto=producto, comentario=comentario,
                  foto=nombre_foto, foto_b64=foto_b64 or None, usuario=usuario,
                  fecha=fecha, semana=semana_actual())
    db.session.add(rep)
    db.session.commit()
    return jsonify({"ok": True, "id": rep.id})

@app.route("/api/reportes")
def ver_reportes():
    tienda = request.args.get("tienda", "")
    semana = request.args.get("semana", semana_actual())
    q = Reporte.query.filter_by(semana=semana)
    if tienda:
        q = q.filter_by(tienda=tienda)
    rows = q.order_by(Reporte.id.desc()).all()
    return jsonify([{"id": r.id, "tienda": r.tienda, "producto": r.producto,
                     "comentario": r.comentario, "foto": r.foto,
                     "foto_b64": r.foto_b64 or "", "usuario": r.usuario, "fecha": r.fecha} for r in rows])

@app.route("/api/cerrar-semana", methods=["POST"])
def cerrar_semana():
    semana = semana_actual()
    reportes = Reporte.query.filter_by(semana=semana).all()

    # Construir mapa de reportes: {(tienda, producto): reporte}
    mapa = {}
    for r in reportes:
        key = (r.tienda, r.producto)
        if key not in mapa or r.foto:  # preferir el que tiene foto
            mapa[key] = r

    # Semana anterior para detectar reincidentes
    semana_ant = Diferencia.query.filter(Diferencia.semana < semana).order_by(Diferencia.semana.desc()).with_entities(Diferencia.semana).first()
    semana_anterior = semana_ant[0] if semana_ant else None
    difs_anteriores = set()
    if semana_anterior:
        for d in Diferencia.query.filter_by(semana=semana_anterior).all():
            if d.estado != "OK":
                difs_anteriores.add((d.tienda, d.producto))

    # Mapa de validación IA: reporte_id → estado IA
    validaciones_ia = {v.reporte_id: v.estado for v in ValidacionIA.query.filter_by(semana=semana).all()}

    # Calcular diferencias
    inventario = Inventario.query.all()
    Diferencia.query.filter_by(semana=semana).delete()
    for inv in inventario:
        if inv.cantidad <= 0:
            continue
        key = (inv.tienda, inv.producto)
        rep = mapa.get(key)
        estado_ia = validaciones_ia.get(rep.id) if rep else None
        if rep and rep.foto and estado_ia == "RECHAZADO":
            # Foto rechazada por IA → cuenta como diferencia
            estado = "SIN_FOTO"
            comentario = "Foto rechazada por validación IA"
        elif rep and rep.foto and estado_ia == "REVISAR":
            # Pendiente de revisión → justificación
            estado = "CON_JUSTIFICACION"
            comentario = "Foto pendiente de revisión IA"
        elif rep and rep.foto:
            estado = "OK"
            comentario = rep.comentario or ""
        elif rep and rep.comentario:
            estado = "CON_JUSTIFICACION"
            comentario = rep.comentario
        else:
            estado = "SIN_FOTO"
            comentario = ""
        db.session.add(Diferencia(semana=semana, tienda=inv.tienda, producto=inv.producto,
                                   cantidad=inv.cantidad, estado=estado, comentario=comentario))
    db.session.commit()

    # Generar Excel
    wb = openpyxl.Workbook()

    # Estilos
    fill_rojo     = PatternFill("solid", fgColor="FFD7D7")
    fill_amarillo = PatternFill("solid", fgColor="FFF3CD")
    fill_verde    = PatternFill("solid", fgColor="D4EDDA")
    fill_header   = PatternFill("solid", fgColor="1E2230")
    fill_reinc    = PatternFill("solid", fgColor="FFB3B3")
    font_header   = Font(bold=True, color="00BF8F", size=11)
    font_reinc    = Font(bold=True, color="C0392B")
    thin          = Side(style="thin", color="CCCCCC")
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)
    center        = Alignment(horizontal="center", vertical="center")

    def estilo_header(ws, cols):
        for cell in ws[1]:
            cell.fill   = fill_header
            cell.font   = font_header
            cell.alignment = center
            cell.border = border
        for i, w in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    rango_archivo, rango_legible = rango_semana_actual()

    # Hoja 1 — Diferencias actuales
    ws1 = wb.active
    ws1.title = f"Diferencias {rango_legible}"
    ws1.append(["Tienda", "Producto", "Stock sistema", "Estado", "Comentario", "Reincidente"])
    estilo_header(ws1, [35, 40, 14, 18, 40, 12])

    difs = Diferencia.query.filter_by(semana=semana).filter(Diferencia.estado != "OK").order_by(Diferencia.tienda, Diferencia.producto).all()
    for d in difs:
        reinc = "SI" if (d.tienda, d.producto) in difs_anteriores else "No"
        estado_txt = "Sin foto" if d.estado == "SIN_FOTO" else "Con justificación"
        row = [d.tienda, d.producto, d.cantidad, estado_txt, d.comentario or "", reinc]
        ws1.append(row)
        r = ws1.max_row
        fill = fill_rojo if d.estado == "SIN_FOTO" else fill_amarillo
        for c in range(1, 7):
            ws1.cell(r, c).fill   = fill_reinc if reinc == "SI" else fill
            ws1.cell(r, c).border = border
            ws1.cell(r, c).alignment = Alignment(vertical="center", wrap_text=True)
        if reinc == "SI":
            ws1.cell(r, 6).font = font_reinc

    # Hoja 2 — Solo reincidentes
    ws2 = wb.create_sheet("Reincidentes")
    ws2.append(["Tienda", "Producto", "Stock sistema", "Estado", "Comentario"])
    estilo_header(ws2, [35, 40, 14, 18, 40])
    for d in difs:
        if (d.tienda, d.producto) in difs_anteriores:
            estado_txt = "Sin foto" if d.estado == "SIN_FOTO" else "Con justificación"
            ws2.append([d.tienda, d.producto, d.cantidad, estado_txt, d.comentario or ""])
            r = ws2.max_row
            for c in range(1, 6):
                ws2.cell(r, c).fill   = fill_reinc
                ws2.cell(r, c).font   = font_reinc
                ws2.cell(r, c).border = border

    # Hoja 3 — Todo OK
    ws3 = wb.create_sheet("Exhibidos OK")
    ws3.append(["Tienda", "Producto", "Stock sistema", "Comentario"])
    estilo_header(ws3, [35, 40, 14, 40])
    for d in Diferencia.query.filter_by(semana=semana, estado="OK").order_by(Diferencia.tienda).all():
        ws3.append([d.tienda, d.producto, d.cantidad, d.comentario or ""])
        r = ws3.max_row
        for c in range(1, 5):
            ws3.cell(r, c).fill   = fill_verde
            ws3.cell(r, c).border = border

    # Hoja 4 — Historial matricial: todas las tiendas, ordenado por total de diferencias
    inventario_actual = {(inv.tienda, inv.producto): inv.cantidad for inv in Inventario.query.all()}

    semanas_arch = sorted(set(
        d.semana for d in Diferencia.query.with_entities(Diferencia.semana).distinct().all()
    ))

    # {(tienda, producto): {semana: 1/0}}
    matriz = {}
    for d in Diferencia.query.order_by(Diferencia.semana).all():
        key = (d.tienda, d.producto)
        if key not in matriz:
            matriz[key] = {}
        if d.estado != "OK":
            matriz[key][d.semana] = 1
        elif d.semana not in matriz[key]:
            matriz[key][d.semana] = 0

    fill_uno        = PatternFill("solid", fgColor="C62828")
    fill_cero       = PatternFill("solid", fgColor="1B5E20")
    fill_vacio      = PatternFill("solid", fgColor="2a2a2a")
    font_blanco     = Font(bold=True, color="FFFFFF", size=10)
    font_header_col = Font(bold=True, color="FFFFFF", size=9)
    fill_header_col = PatternFill("solid", fgColor="1e4060")
    fill_total_alto = PatternFill("solid", fgColor="5C0000")

    ws4 = wb.create_sheet("Historial diferencias")

    cabeceras = ["Tienda", "Producto", "Stock"]
    for sem in semanas_arch:
        _, leg = rango_de_codigo(sem)
        cabeceras.append(leg)
    cabeceras.append("Total")
    cabeceras.append("Comentario semana actual")
    ws4.append(cabeceras)

    anchos = [32, 40, 8] + [16] * len(semanas_arch) + [8, 45]
    for i, ancho in enumerate(anchos, 1):
        ws4.column_dimensions[ws4.cell(1, i).column_letter].width = ancho
        ws4.cell(1, i).fill = fill_header_col
        ws4.cell(1, i).font = font_header_col
        ws4.cell(1, i).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws4.row_dimensions[1].height = 40

    # Comentarios de la semana que se está cerrando — directo del mapa de reportes
    comentarios_semana = {
        key: (rep.comentario or "")
        for key, rep in mapa.items()
        if rep.comentario
    }

    # Calcular total por fila y ordenar de mayor a menor
    filas_mat = []
    for (tienda, producto), sem_map in matriz.items():
        stock = inventario_actual.get((tienda, producto), "")
        vals = [sem_map.get(s, None) for s in semanas_arch]
        total = sum(1 for v in vals if v == 1)
        comentario = comentarios_semana.get((tienda, producto), "")
        filas_mat.append((tienda, producto, stock, vals, total, comentario))
    filas_mat.sort(key=lambda x: x[4], reverse=True)

    for tienda, producto, stock, vals, total, comentario in filas_mat:
        fila = [tienda, producto, stock] + [""] * len(semanas_arch) + [total, comentario]
        ws4.append(fila)
        r = ws4.max_row
        # Tienda / Producto / Stock
        for c in range(1, 4):
            ws4.cell(r, c).border = border
            ws4.cell(r, c).alignment = Alignment(vertical="center")
        # Celdas de semana
        for i, v in enumerate(vals):
            c = 4 + i
            ws4.cell(r, c).border = border
            ws4.cell(r, c).alignment = Alignment(horizontal="center", vertical="center")
            if v == 1:
                ws4.cell(r, c).fill = fill_uno
                ws4.cell(r, c).font = font_blanco
                ws4.cell(r, c).value = "1"
            elif v == 0:
                ws4.cell(r, c).fill = fill_cero
                ws4.cell(r, c).font = font_blanco
                ws4.cell(r, c).value = "0"
            else:
                ws4.cell(r, c).fill = fill_vacio
        # Total
        c_total = 4 + len(semanas_arch)
        ws4.cell(r, c_total).border = border
        ws4.cell(r, c_total).alignment = Alignment(horizontal="center", vertical="center")
        if total >= 3:
            ws4.cell(r, c_total).fill = fill_total_alto
            ws4.cell(r, c_total).font = Font(bold=True, color="FFFFFF")
        # Comentario
        c_com = c_total + 1
        ws4.cell(r, c_com).border = border
        ws4.cell(r, c_com).alignment = Alignment(vertical="center", wrap_text=True)

    # Hoja 5 — Tiendas sin visita
    todas_tiendas = set(inv.tienda for inv in Inventario.query.all())
    tiendas_con_reporte = set(r.tienda for r in reportes)
    tiendas_sin_visita = sorted(todas_tiendas - tiendas_con_reporte)

    fill_naranja = PatternFill("solid", fgColor="FFE0B2")
    font_naranja = Font(bold=True, color="BF360C")

    ws5 = wb.create_sheet("Tiendas sin visita")
    ws5.append(["Tienda", "Productos en sistema", "Observación"])
    estilo_header(ws5, [40, 20, 40])
    for t in tiendas_sin_visita:
        prods = Inventario.query.filter_by(tienda=t).count()
        ws5.append([t, prods, "Ningún display reportó esta tienda en la semana"])
        r = ws5.max_row
        for c in range(1, 4):
            ws5.cell(r, c).fill   = fill_naranja
            ws5.cell(r, c).font   = font_naranja
            ws5.cell(r, c).border = border

    # Hoja 6 — Resumen por tienda con % cumplimiento
    ws6 = wb.create_sheet("Resumen por tienda")
    ws6.append(["Tienda", "Total productos", "Con foto", "Sin foto", "% Cumplimiento"])
    estilo_header(ws6, [40, 16, 12, 12, 16])

    resumen_tiendas = {}
    for inv in Inventario.query.all():
        if inv.cantidad <= 0:
            continue
        t = inv.tienda
        if t not in resumen_tiendas:
            resumen_tiendas[t] = {"total": 0, "ok": 0}
        resumen_tiendas[t]["total"] += 1

    for d in Diferencia.query.filter_by(semana=semana, estado="OK").all():
        if d.tienda in resumen_tiendas:
            resumen_tiendas[d.tienda]["ok"] += 1

    filas_resumen = []
    for t, datos in resumen_tiendas.items():
        pct = round(datos["ok"] / datos["total"] * 100) if datos["total"] else 0
        filas_resumen.append((t, datos["total"], datos["ok"], datos["total"] - datos["ok"], pct))
    filas_resumen.sort(key=lambda x: x[4])  # menor % primero

    for t, total, ok, sin_foto, pct in filas_resumen:
        ws6.append([t, total, ok, sin_foto, f"{pct}%"])
        r = ws6.max_row
        fill = fill_verde if pct >= 80 else (fill_amarillo if pct >= 50 else fill_rojo)
        for c in range(1, 6):
            ws6.cell(r, c).fill   = fill
            ws6.cell(r, c).border = border
            ws6.cell(r, c).alignment = Alignment(vertical="center")

    # Hoja 7 — Resumen por display
    ws7 = wb.create_sheet("Resumen por display")
    ws7.append(["Display", "Tiendas visitadas", "Productos reportados", "Con foto", "Última actividad"])
    estilo_header(ws7, [25, 18, 20, 12, 20])

    resumen_display = {}
    for rep in reportes:
        u = rep.usuario or "Desconocido"
        if u not in resumen_display:
            resumen_display[u] = {"tiendas": set(), "total": 0, "fotos": 0, "ultima": ""}
        resumen_display[u]["tiendas"].add(rep.tienda)
        resumen_display[u]["total"] += 1
        if rep.foto:
            resumen_display[u]["fotos"] += 1
        if rep.fecha > resumen_display[u]["ultima"]:
            resumen_display[u]["ultima"] = rep.fecha

    for disp, datos in sorted(resumen_display.items()):
        fecha_fmt = datos["ultima"][:8]
        if len(fecha_fmt) == 8:
            fecha_fmt = f"{fecha_fmt[6:8]}/{fecha_fmt[4:6]}/{fecha_fmt[:4]}"
        ws7.append([disp, len(datos["tiendas"]), datos["total"], datos["fotos"], fecha_fmt])
        r = ws7.max_row
        for c in range(1, 6):
            ws7.cell(r, c).fill   = fill_verde
            ws7.cell(r, c).border = border
            ws7.cell(r, c).alignment = Alignment(vertical="center")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Recopilar fotos antes de borrar reportes
    fotos = {}  # {(tienda, producto): foto_b64_str}
    for (tienda, producto), rep in mapa.items():
        if rep.foto_b64:
            fotos[(tienda, producto)] = rep.foto_b64

    # Borrar reportes de la semana cerrada para empezar semana nueva limpia
    Reporte.query.filter_by(semana=semana).delete()
    db.session.commit()

    # Construir ZIP: Excel + carpetas de fotos por tienda
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # Excel dentro del ZIP
        zf.writestr(f"Diferencias_{rango_archivo}.xlsx", buf.getvalue())

        # Fotos organizadas en carpetas: Fotos_<rango>/<Tienda>/<Producto>.jpg
        for (tienda, producto), foto_b64 in fotos.items():
            try:
                # El frontend puede enviar "data:image/jpeg;base64,..." o solo base64 puro
                raw = foto_b64.split(",", 1)[1] if "," in foto_b64 else foto_b64
                foto_bytes = base64.b64decode(raw)
                tienda_safe   = re.sub(r'[\\/:*?"<>|]', "_", tienda).strip()
                producto_safe = re.sub(r'[\\/:*?"<>|]', "_", producto).strip()[:60]
                path = f"Fotos_{rango_archivo}/{tienda_safe}/{producto_safe}.jpg"
                zf.writestr(path, foto_bytes)
            except Exception:
                pass  # foto corrupta o vacía, se omite

    zip_buf.seek(0)
    nombre_zip = f"Cierre_{rango_archivo}.zip"
    return send_file(zip_buf, as_attachment=True, download_name=nombre_zip,
                     mimetype="application/zip")

@app.route("/api/limpiar-reportes", methods=["POST"])
def limpiar_reportes():
    data = request.json or {}
    if data.get("rol") != "supervisor":
        return jsonify({"error": "No autorizado"}), 403
    count = Reporte.query.count()
    Reporte.query.delete()
    db.session.commit()
    return jsonify({"ok": True, "borrados": count})

@app.route("/api/limpiar-historial", methods=["POST"])
def limpiar_historial():
    data = request.json or {}
    if data.get("rol") != "supervisor":
        return jsonify({"error": "No autorizado"}), 403
    count = Diferencia.query.count()
    Diferencia.query.delete()
    db.session.commit()
    return jsonify({"ok": True, "borrados": count})

@app.route("/api/limpiar-todo", methods=["POST"])
def limpiar_todo():
    data = request.json or {}
    if data.get("rol") != "supervisor":
        return jsonify({"error": "No autorizado"}), 403
    r = Reporte.query.count()
    d = Diferencia.query.count()
    Reporte.query.delete()
    Diferencia.query.delete()
    db.session.commit()
    return jsonify({"ok": True, "reportes_borrados": r, "historial_borrado": d})

@app.route("/api/reporte/<int:reporte_id>/foto")
def ver_foto_reporte(reporte_id):
    rep = Reporte.query.get_or_404(reporte_id)
    return jsonify({"foto_b64": rep.foto_b64 or ""})

@app.route("/api/diferencias-semana")
def diferencias_semana():
    semana = semana_actual()
    reportes = Reporte.query.filter_by(semana=semana).all()
    mapa = {}
    for r in reportes:
        key = (r.tienda, r.producto)
        if key not in mapa or r.foto:
            mapa[key] = r
    inventario = Inventario.query.filter(Inventario.cantidad > 0).all()
    tiendas_con_reporte = set(r.tienda for r in reportes)
    todas_tiendas = set(inv.tienda for inv in Inventario.query.all())
    resultado = []
    for inv in inventario:
        key = (inv.tienda, inv.producto)
        if inv.cantidad <= 0:
            continue  # sin stock en sistema, no es diferencia
        rep = mapa.get(key)
        if rep and rep.foto:
            estado = "OK"
        elif rep and rep.comentario:
            estado = "CON_JUSTIFICACION"
        else:
            estado = "SIN_FOTO"
        resultado.append({"tienda": inv.tienda, "producto": inv.producto,
                          "cantidad": inv.cantidad, "estado": estado,
                          "comentario": rep.comentario if rep else ""})
    sin_visita = sorted(todas_tiendas - tiendas_con_reporte)
    return jsonify({"diferencias": resultado, "tiendas_sin_visita": sin_visita})

@app.route("/api/reportes/<int:reporte_id>", methods=["DELETE"])
def eliminar_reporte(reporte_id):
    rep = Reporte.query.get_or_404(reporte_id)
    db.session.delete(rep)
    db.session.commit()
    return jsonify({"ok": True})

# ═══════════════════════════════════════════════════════════════════════════════
# VALIDACIÓN IA — Fase 1: encolar y consultar progreso
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/validacion/encolar", methods=["POST"])
def validacion_encolar():
    """Crea registros PENDIENTE en ValidacionIA para todos los reportes con foto de la semana actual."""
    data = request.json or {}
    if data.get("rol") != "supervisor":
        return jsonify({"error": "No autorizado"}), 403

    semana = semana_actual()
    reportes = Reporte.query.filter_by(semana=semana).all()

    encolados = 0
    for rep in reportes:
        if not rep.foto_b64:
            continue
        # Evitar duplicados
        existe = ValidacionIA.query.filter_by(reporte_id=rep.id).first()
        if existe:
            continue
        # Buscar marca en inventario (por ahora vacía, se enriquece luego)
        val = ValidacionIA(
            reporte_id   = rep.id,
            semana       = semana,
            tienda       = rep.tienda,
            producto     = rep.producto,
            marca        = "",
            estado       = "PENDIENTE",
            creado_en    = datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        )
        db.session.add(val)
        encolados += 1

    db.session.commit()
    return jsonify({"ok": True, "encolados": encolados, "semana": semana})


@app.route("/api/validacion/progreso")
def validacion_progreso():
    """Retorna el progreso de validación IA de la semana actual."""
    semana = request.args.get("semana", semana_actual())
    total      = ValidacionIA.query.filter_by(semana=semana).count()
    pendientes = ValidacionIA.query.filter_by(semana=semana, estado="PENDIENTE").count()
    procesando = ValidacionIA.query.filter_by(semana=semana, estado="PROCESANDO").count()
    aprobados  = ValidacionIA.query.filter_by(semana=semana, estado="APROBADO").count()
    rechazados = ValidacionIA.query.filter_by(semana=semana, estado="RECHAZADO").count()
    revisar    = ValidacionIA.query.filter_by(semana=semana, estado="REVISAR").count()
    errores    = ValidacionIA.query.filter_by(semana=semana, estado="ERROR").count()

    costo_total = db.session.query(db.func.sum(ValidacionIA.costo_usd))\
        .filter_by(semana=semana).scalar() or 0.0

    return jsonify({
        "semana": semana,
        "total": total,
        "pendientes": pendientes,
        "procesando": procesando,
        "aprobados": aprobados,
        "rechazados": rechazados,
        "revisar": revisar,
        "errores": errores,
        "procesadas": aprobados + rechazados + revisar + errores,
        "costo_usd": round(costo_total, 4),
    })


@app.route("/api/validacion/resultados")
def validacion_resultados():
    """Lista de validaciones con filtro por estado para revisión manual."""
    semana = request.args.get("semana", semana_actual())
    estado = request.args.get("estado", "")  # RECHAZADO, REVISAR, etc.

    q = ValidacionIA.query.filter_by(semana=semana)
    if estado:
        q = q.filter_by(estado=estado)
    q = q.order_by(ValidacionIA.estado, ValidacionIA.tienda)

    rows = q.all()
    resultado = []
    for v in rows:
        foto_b64 = ""
        if v.reporte_id:
            rep = Reporte.query.get(v.reporte_id)
            if rep:
                foto_b64 = rep.foto_b64 or ""
        resultado.append({
            "id": v.id,
            "tienda": v.tienda,
            "producto": v.producto,
            "estado": v.estado,
            "confianza": v.confianza or "",
            "motivo": v.motivo or "",
            "costo_usd": v.costo_usd or 0,
            "foto_b64": foto_b64,
        })
    return jsonify(resultado)


# ═══════════════════════════════════════════════════════════════════════════════
# VALIDACIÓN IA — Fase 2: worker que llama a Claude Haiku
# ═══════════════════════════════════════════════════════════════════════════════

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")

_worker_lock = threading.Lock()
_worker_running = False


def _procesar_validaciones():
    """Procesa en background todos los registros PENDIENTE llamando a Claude Haiku."""
    global _worker_running
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

        with app.app_context():
            while True:
                val = ValidacionIA.query.filter_by(estado="PENDIENTE").first()
                if not val:
                    break

                # Marcar como PROCESANDO para evitar doble proceso
                val.estado = "PROCESANDO"
                val.intentos = (val.intentos or 0) + 1
                db.session.commit()

                try:
                    # Obtener foto del reporte
                    rep = Reporte.query.get(val.reporte_id) if val.reporte_id else None
                    if not rep or not rep.foto_b64:
                        val.estado = "ERROR"
                        val.motivo = "Sin foto disponible"
                        val.procesado_en = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        db.session.commit()
                        continue

                    # Preparar imagen en base64
                    foto_raw = rep.foto_b64
                    if "," in foto_raw:
                        foto_raw = foto_raw.split(",", 1)[1]

                    prompt = f"""Eres un auditor de displays de productos en supermercados.
Analiza esta foto y determina si muestra correctamente el producto "{val.producto}" exhibido en tienda.

Responde EXACTAMENTE en este formato (sin texto adicional):
ESTADO: [APROBADO|RECHAZADO|REVISAR]
CONFIANZA: [alta|media|baja]
MOTIVO: [una sola oración explicando tu decisión]

Criterios:
- APROBADO: la foto muestra claramente el producto correcto bien exhibido
- RECHAZADO: foto incorrecta (selfie, piso, producto diferente, foto de otra cosa)
- REVISAR: foto borrosa, muy oscura, o producto parcialmente visible — requiere revisión humana"""

                    response = client.messages.create(
                        model="claude-haiku-4-5-20251001",
                        max_tokens=150,
                        messages=[{
                            "role": "user",
                            "content": [
                                {
                                    "type": "image",
                                    "source": {
                                        "type": "base64",
                                        "media_type": "image/jpeg",
                                        "data": foto_raw,
                                    },
                                },
                                {"type": "text", "text": prompt},
                            ],
                        }],
                    )

                    texto = response.content[0].text.strip()
                    tokens_in  = response.usage.input_tokens
                    tokens_out = response.usage.output_tokens
                    # Precio Haiku: $0.80/MTok input, $4.00/MTok output
                    costo = (tokens_in * 0.80 + tokens_out * 4.00) / 1_000_000

                    # Parsear respuesta
                    estado_ia   = "REVISAR"
                    confianza   = "baja"
                    motivo      = texto

                    for linea in texto.splitlines():
                        linea = linea.strip()
                        if linea.startswith("ESTADO:"):
                            v = linea.split(":", 1)[1].strip().upper()
                            if v in ("APROBADO", "RECHAZADO", "REVISAR"):
                                estado_ia = v
                        elif linea.startswith("CONFIANZA:"):
                            c = linea.split(":", 1)[1].strip().lower()
                            if c in ("alta", "media", "baja"):
                                confianza = c
                        elif linea.startswith("MOTIVO:"):
                            motivo = linea.split(":", 1)[1].strip()

                    val.estado       = estado_ia
                    val.confianza    = confianza
                    val.motivo       = motivo
                    val.tokens_input = tokens_in
                    val.tokens_output= tokens_out
                    val.costo_usd    = costo
                    val.procesado_en = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    db.session.commit()

                except Exception as e:
                    val.estado    = "ERROR"
                    val.motivo    = str(e)[:300]
                    val.procesado_en = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    db.session.commit()
    finally:
        with _worker_lock:
            _worker_running = False


@app.route("/api/validacion/<int:val_id>/decision", methods=["POST"])
def validacion_decision(val_id):
    """Supervisor aprueba o deniega manualmente una validación REVISAR o RECHAZADO."""
    data = request.json or {}
    if data.get("rol") != "supervisor":
        return jsonify({"error": "No autorizado"}), 403
    decision = data.get("decision", "").upper()
    if decision not in ("APROBADO", "RECHAZADO"):
        return jsonify({"error": "decision debe ser APROBADO o RECHAZADO"}), 400
    val = ValidacionIA.query.get_or_404(val_id)
    val.estado = decision
    val.motivo = (val.motivo or "") + f" [Revisado manualmente por supervisor]"
    val.procesado_en = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    db.session.commit()
    return jsonify({"ok": True, "estado": val.estado})


@app.route("/api/validacion/procesar", methods=["POST"])
def validacion_procesar():
    """Lanza el worker en background para procesar fotos PENDIENTE con Claude Haiku."""
    global _worker_running
    if not ANTHROPIC_API_KEY:
        return jsonify({"error": "ANTHROPIC_API_KEY no configurada"}), 500

    data = request.json or {}
    if data.get("rol") != "supervisor":
        return jsonify({"error": "No autorizado"}), 403

    pendientes = ValidacionIA.query.filter_by(estado="PENDIENTE").count()
    if pendientes == 0:
        return jsonify({"ok": True, "mensaje": "No hay fotos pendientes", "iniciado": False})

    with _worker_lock:
        if _worker_running:
            return jsonify({"ok": True, "mensaje": "Worker ya en ejecución", "iniciado": False})
        _worker_running = True

    t = threading.Thread(target=_procesar_validaciones, daemon=True)
    t.start()
    return jsonify({"ok": True, "mensaje": f"Procesando {pendientes} fotos en background", "iniciado": True})


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
